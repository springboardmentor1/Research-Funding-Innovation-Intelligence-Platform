"""
Excel Report Generator.

Uses openpyxl to create styled Excel workbooks for:
  - Funding opportunities
  - Patent analysis
  - Research trends
"""

import io
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Styling ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
TITLE_FONT = Font(name="Calibri", bold=True, color="6366F1", size=14)


def _style_header(ws, row_num: int, col_count: int):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data_row(ws, row_num: int, col_count: int, alt: bool = False):
    """Apply data row styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def _auto_width(ws, col_count: int, max_width: int = 40):
    """Auto-adjust column widths based on content."""
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    for col_idx in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _add_title(ws, title: str, col_count: int):
    """Add a title row and timestamp."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ts_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ts_cell.font = Font(name="Calibri", italic=True, color="666666", size=9)
    ts_cell.alignment = Alignment(horizontal="center")


# ── Report Generators ─────────────────────────────────────────────────────────


def generate_funding_excel(data: List[Dict[str, Any]]) -> bytes:
    """Generate a Funding Opportunities Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Funding Opportunities"

    headers = ["Grant Name", "Organization", "Area", "Amount", "Deadline", "Country", "Description"]
    col_count = len(headers)

    _add_title(ws, "Funding Opportunities Report", col_count)

    # Header row
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, header_row, col_count)

    # Data rows
    for idx, row in enumerate(data):
        r = header_row + 1 + idx
        ws.cell(row=r, column=1, value=str(row.get("Grant", "")))
        ws.cell(row=r, column=2, value=str(row.get("Organization", "")))
        ws.cell(row=r, column=3, value=str(row.get("Area", "")))
        ws.cell(row=r, column=4, value=str(row.get("Amount", "")))
        ws.cell(row=r, column=5, value=str(row.get("Deadline", "")))
        ws.cell(row=r, column=6, value=str(row.get("Country", "")))
        ws.cell(row=r, column=7, value=str(row.get("Description", ""))[:200])
        _style_data_row(ws, r, col_count, alt=(idx % 2 == 1))

    _auto_width(ws, col_count)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_patent_excel(landscape: Dict[str, Any], trends: Dict[str, Any]) -> bytes:
    """Generate a Patent Analysis Excel report."""
    wb = Workbook()

    # Sheet 1: By Technology
    ws1 = wb.active
    ws1.title = "By Technology"
    headers1 = ["Technology", "Count"]
    _add_title(ws1, "Patent Analysis — By Technology", len(headers1))
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=i, value=h)
    _style_header(ws1, 4, len(headers1))
    for idx, item in enumerate(landscape.get("by_technology", [])):
        r = 5 + idx
        ws1.cell(row=r, column=1, value=str(item.get("Technology", "")))
        ws1.cell(row=r, column=2, value=int(item.get("count", 0)))
        _style_data_row(ws1, r, len(headers1), alt=(idx % 2 == 1))
    _auto_width(ws1, len(headers1))

    # Sheet 2: Trends
    ws2 = wb.create_sheet("Trends")
    headers2 = ["Year", "Patents", "Growth %"]
    _add_title(ws2, "Patent Filing Trends", len(headers2))
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=i, value=h)
    _style_header(ws2, 4, len(headers2))
    for idx, t in enumerate(trends.get("trends", [])):
        r = 5 + idx
        ws2.cell(row=r, column=1, value=int(t.get("year", 0)))
        ws2.cell(row=r, column=2, value=int(t.get("count", 0)))
        ws2.cell(row=r, column=3, value=float(t.get("growth_pct", 0)))
        _style_data_row(ws2, r, len(headers2), alt=(idx % 2 == 1))
    _auto_width(ws2, len(headers2))

    # Sheet 3: Top Assignees
    ws3 = wb.create_sheet("Top Assignees")
    headers3 = ["Assignee", "Count"]
    _add_title(ws3, "Top Patent Assignees", len(headers3))
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=4, column=i, value=h)
    _style_header(ws3, 4, len(headers3))
    for idx, item in enumerate(landscape.get("by_assignee", [])):
        r = 5 + idx
        ws3.cell(row=r, column=1, value=str(item.get("Assignee", "")))
        ws3.cell(row=r, column=2, value=int(item.get("count", 0)))
        _style_data_row(ws3, r, len(headers3), alt=(idx % 2 == 1))
    _auto_width(ws3, len(headers3))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_research_excel(trends: Dict[str, Any], keywords: List[Dict]) -> bytes:
    """Generate a Research Trends Excel report."""
    wb = Workbook()

    # Sheet 1: Publication Trends
    ws1 = wb.active
    ws1.title = "Publication Trends"
    headers1 = ["Year", "Papers", "Growth %"]
    _add_title(ws1, "Research Publication Trends", len(headers1))
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=i, value=h)
    _style_header(ws1, 4, len(headers1))
    for idx, t in enumerate(trends.get("trends", [])):
        r = 5 + idx
        ws1.cell(row=r, column=1, value=int(t.get("year", 0)))
        ws1.cell(row=r, column=2, value=int(t.get("count", 0)))
        ws1.cell(row=r, column=3, value=float(t.get("growth_pct", 0)))
        _style_data_row(ws1, r, len(headers1), alt=(idx % 2 == 1))
    _auto_width(ws1, len(headers1))

    # Sheet 2: Top Keywords
    ws2 = wb.create_sheet("Top Keywords")
    headers2 = ["Keyword", "Count"]
    _add_title(ws2, "Top Research Keywords", len(headers2))
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=i, value=h)
    _style_header(ws2, 4, len(headers2))
    for idx, kw in enumerate(keywords):
        r = 5 + idx
        ws2.cell(row=r, column=1, value=str(kw.get("keyword", "")))
        ws2.cell(row=r, column=2, value=int(kw.get("count", 0)))
        _style_data_row(ws2, r, len(headers2), alt=(idx % 2 == 1))
    _auto_width(ws2, len(headers2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_innovation_excel(scores: Dict[str, Any], top_patents: List[Dict]) -> bytes:
    """Generate an Innovation Intelligence Excel report."""
    wb = Workbook()

    # Sheet 1: Top Innovations
    ws1 = wb.active
    ws1.title = "Top Innovations"
    headers1 = ["Title", "Technology", "Innovation Score", "Year", "Novelty", "Strength", "Maturity", "Market", "Funding"]
    _add_title(ws1, "Innovation Intelligence — Top Innovations", len(headers1))
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=i, value=h)
    _style_header(ws1, 4, len(headers1))

    for idx, p in enumerate(top_patents):
        r = 5 + idx
        breakdown = p.get("breakdown", {})
        ws1.cell(row=r, column=1, value=str(p.get("title", ""))[:120])
        ws1.cell(row=r, column=2, value=str(p.get("technology", "")))
        ws1.cell(row=r, column=3, value=float(p.get("innovation_score", 0)))
        ws1.cell(row=r, column=4, value=int(p.get("year", 0)))
        ws1.cell(row=r, column=5, value=float(breakdown.get("research_novelty", 0)))
        ws1.cell(row=r, column=6, value=float(breakdown.get("patent_strength", 0)))
        ws1.cell(row=r, column=7, value=float(breakdown.get("tech_maturity", 0)))
        ws1.cell(row=r, column=8, value=float(breakdown.get("market_potential", 0)))
        ws1.cell(row=r, column=9, value=float(breakdown.get("funding_relevance", 0)))
        _style_data_row(ws1, r, len(headers1), alt=(idx % 2 == 1))
    _auto_width(ws1, len(headers1))

    # Sheet 2: Score Distribution
    ws2 = wb.create_sheet("Score Distribution")
    headers2 = ["Score Range", "Count"]
    _add_title(ws2, "Innovation Score Distribution", len(headers2))
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=i, value=h)
    _style_header(ws2, 4, len(headers2))
    for idx, d in enumerate(scores.get("distribution", [])):
        r = 5 + idx
        ws2.cell(row=r, column=1, value=str(d.get("range", "")))
        ws2.cell(row=r, column=2, value=int(d.get("count", 0)))
        _style_data_row(ws2, r, len(headers2), alt=(idx % 2 == 1))
    _auto_width(ws2, len(headers2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_commercialization_excel(data: Dict[str, Any]) -> bytes:
    """Generate a Commercialization Recommendations Excel report."""
    wb = Workbook()

    # Sheet 1: Top Commercializable
    ws1 = wb.active
    ws1.title = "Commercializable Patents"
    headers1 = ["Title", "Technology", "Innovation Score", "Action", "Confidence", "Time to Market"]
    _add_title(ws1, "Commercialization Recommendations", len(headers1))
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=i, value=h)
    _style_header(ws1, 4, len(headers1))

    top = data.get("top_commercializable", [])
    for idx, p in enumerate(top):
        r = 5 + idx
        rec = p.get("recommendation", {})
        ws1.cell(row=r, column=1, value=str(p.get("title", ""))[:120])
        ws1.cell(row=r, column=2, value=str(p.get("technology", "")))
        ws1.cell(row=r, column=3, value=float(p.get("innovation_score", 0)))
        ws1.cell(row=r, column=4, value=str(rec.get("action", "")))
        ws1.cell(row=r, column=5, value=str(rec.get("confidence", "")))
        ws1.cell(row=r, column=6, value=str(rec.get("time_to_market", "")))
        _style_data_row(ws1, r, len(headers1), alt=(idx % 2 == 1))
    _auto_width(ws1, len(headers1))

    # Sheet 2: Action Distribution
    ws2 = wb.create_sheet("Action Distribution")
    headers2 = ["Recommended Action", "Count", "Percentage"]
    _add_title(ws2, "Commercialization Action Distribution", len(headers2))
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=i, value=h)
    _style_header(ws2, 4, len(headers2))
    for idx, d in enumerate(data.get("distribution", [])):
        r = 5 + idx
        ws2.cell(row=r, column=1, value=str(d.get("action", "")))
        ws2.cell(row=r, column=2, value=int(d.get("count", 0)))
        ws2.cell(row=r, column=3, value=f"{d.get('percentage', 0)}%")
        _style_data_row(ws2, r, len(headers2), alt=(idx % 2 == 1))
    _auto_width(ws2, len(headers2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

