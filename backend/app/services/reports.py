"""
Report generation: PDF and Excel export (Module 11).

Builds downloadable reports IN MEMORY and streams them back, rather than
writing temp files on disk. A BytesIO buffer means nothing to clean up and
no race between concurrent requests writing the same path.

Two formats because they serve different readers:
    Excel - an analyst who wants the numbers to pivot and re-chart
    PDF   - a stakeholder who wants a fixed document to read or print
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from sqlalchemy.orm import Session

from app.models import ResearchProfile
from app.services import analytics, scoring


# ------------------------------------------------------------------ Excel
def _sheet(wb: Workbook, title: str, header: list[str], rows: list[list]):
    ws = wb.create_sheet(title[:31])       # Excel caps sheet names at 31 chars
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4472C4")

    ws.append(header)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill

    for r in rows:
        ws.append(r)

    # autofit-ish: widen each column to its longest value
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None),
                    default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 60)


def build_excel(db: Session, profile: ResearchProfile | None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)      # drop the default empty sheet

    pubs = analytics.publications_per_year(db)
    _sheet(wb, "Publications by Year", ["Year", "Count"],
           [[r["year"], r["count"]] for r in pubs])

    topics = analytics.top_topics(db, 20)
    _sheet(wb, "Top Topics", ["Topic", "Count"],
           [[r["topic"], r["count"]] for r in topics])

    pv = analytics.patent_volume_by_year(db)
    _sheet(wb, "Patent Volume", ["Year", "Count", "Avg Citations"],
           [[r["year"], r["count"], r["avg_citations"]] for r in pv])

    apps = analytics.top_applicants(db, 20)
    _sheet(wb, "Top Applicants", ["Applicant", "Patents"],
           [[r["applicant"], r["count"]] for r in apps])

    cpc = analytics.top_cpc_groups(db, 20)
    _sheet(wb, "Top CPC Groups", ["CPC Group", "Count"],
           [[r["cpc_group"], r["count"]] for r in cpc])

    if profile is not None:
        s = scoring.compute_score(db, profile)
        _sheet(wb, "Innovation Score",
               ["Component", "Value", "Weight", "Contribution"],
               [[k, v["value"], v["weight"], v["contribution"]]
                for k, v in s["components"].items()]
               + [["TOTAL", s["total_score"], 1.0, s["total_score"]]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ PDF
def build_pdf(db: Session, profile: ResearchProfile | None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Research Funding & Innovation Intelligence", styles["Title"]))
    story.append(Paragraph(f"Generated {date.today().isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.6 * cm))

    def table(title, header, rows):
        story.append(Paragraph(title, styles["Heading2"]))
        data = [header] + rows
        t = Table(data, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F2F2F2")]),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5 * cm))

    pv = analytics.patent_volume_by_year(db)
    table("Patent Volume by Year", ["Year", "Count", "Avg Citations"],
          [[r["year"], r["count"], r["avg_citations"]] for r in pv])

    apps = analytics.top_applicants(db, 10)
    table("Top 10 Patent Applicants", ["Applicant", "Patents"],
          [[r["applicant"][:40], r["count"]] for r in apps])

    topics = analytics.top_topics(db, 10)
    table("Top 10 Research Topics", ["Topic", "Count"],
          [[r["topic"][:45], r["count"]] for r in topics])

    if profile is not None:
        s = scoring.compute_score(db, profile)
        table("Innovation Score",
              ["Component", "Value", "Weight", "Contribution"],
              [[k.replace("_", " ").title(), v["value"], v["weight"],
                v["contribution"]] for k, v in s["components"].items()]
              + [["TOTAL", s["total_score"], "", s["total_score"]]])

    doc.build(story)
    return buf.getvalue()
