from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.services import funding_opportunity_service
from app.services.dashboard_service import get_dashboard
from app.services import patent_service
from app.utils.pdf_generator import PDFGenerator
pdf_generator = PDFGenerator()
from app.services.publication_service import (
    get_publication_summary,
    get_yearly_publication_trend,
    get_research_area_trend,
    get_journal_trend,
    get_publications,
)

def generate_dashboard_pdf(
    db: Session,
    user_id: int,
    user_name: str | None = None,
) -> BytesIO:
    """
    Generate a PDF report containing the authenticated
    user's complete dashboard information.

    Parameters:
        db:
            SQLAlchemy database session.

        user_id:
            ID of the authenticated user.

        user_name:
            Name of the authenticated user.

    Returns:
        BytesIO object containing the generated PDF.
    """

    # Get the existing dashboard data.
    dashboard_data = get_dashboard(
        db=db,
        user_id=user_id,
    )

    # Create the PDF generator.
    pdf_generator = PDFGenerator()

    # Generate the dashboard PDF.
    pdf_file = pdf_generator.generate_dashboard_report(
        dashboard_data=dashboard_data,
        user_name=user_name,
    )

    return pdf_file

def generate_publications_pdf(
    db: Session,
    user_id: int,
    user_name: str | None = None,
) -> BytesIO:
    """
    Generate a PDF report containing the authenticated
    user's publication information and analytics.

    Parameters:
        db:
            SQLAlchemy database session.

        user_id:
            ID of the authenticated user.

        user_name:
            Name of the authenticated user.

    Returns:
        BytesIO object containing the generated PDF.
    """

    # --------------------------------------------------------------
    # Publication Summary
    # --------------------------------------------------------------

    summary = get_publication_summary(
        db=db,
        user_id=user_id,
    )

    # --------------------------------------------------------------
    # Publication Trends
    # --------------------------------------------------------------

    yearly_trend = get_yearly_publication_trend(
        db=db,
        user_id=user_id,
    )

    research_area_trend = get_research_area_trend(
        db=db,
        user_id=user_id,
    )

    journal_trend = get_journal_trend(
        db=db,
        user_id=user_id,
    )

    # --------------------------------------------------------------
    # Publication Records
    # --------------------------------------------------------------

    publications = get_publications(
        db=db,
        user_id=user_id,
    )

    # --------------------------------------------------------------
    # Prepare Report Data
    # --------------------------------------------------------------

    report_data = {
        "summary": summary,
        "yearly_trend": yearly_trend,
        "research_area_trend": research_area_trend,
        "journal_trend": journal_trend,
        "publications": publications,
    }

    # --------------------------------------------------------------
    # Generate PDF
    # --------------------------------------------------------------

    pdf_generator = PDFGenerator()

    return pdf_generator.generate_publications_report(
        publication_data=report_data,
        user_name=user_name,
    )

def generate_publications_excel(
    db: Session,
    user_id: int,
    user_name: str | None = None,
) -> BytesIO:
    """
    Generate an Excel report containing the authenticated
    user's publications and publication analytics.
    """

    summary = get_publication_summary(
        db=db,
        user_id=user_id,
    )

    yearly_trend = get_yearly_publication_trend(
        db=db,
        user_id=user_id,
    )

    research_area_trend = get_research_area_trend(
        db=db,
        user_id=user_id,
    )

    journal_trend = get_journal_trend(
        db=db,
        user_id=user_id,
    )

    publications = get_publications(
        db=db,
        user_id=user_id,
    )

    workbook = Workbook()

    # --------------------------------------------------
    # Summary Sheet
    # --------------------------------------------------

    sheet = workbook.active
    sheet.title = "Summary"

    sheet.append([
        "Research Funding & Innovation Intelligence Platform"
    ])

    sheet.append([
        "Publications Report"
    ])

    sheet.append([
        "User",
        user_name or "N/A"
    ])

    sheet.append([])

    sheet.append([
        "Metric",
        "Value"
    ])

    for key, value in summary.items():
        sheet.append([
            str(key).replace("_", " ").title(),
            value,
        ])

    # --------------------------------------------------
    # Yearly Trend
    # --------------------------------------------------

    sheet = workbook.create_sheet("Yearly Trend")

    sheet.append([
        "Year",
        "Publication Count",
    ])

    for item in yearly_trend:
        sheet.append([
            item["year"],
            item["count"],
        ])

    # --------------------------------------------------
    # Research Areas
    # --------------------------------------------------

    sheet = workbook.create_sheet("Research Areas")

    sheet.append([
        "Research Area",
        "Publication Count",
    ])

    for item in research_area_trend:
        sheet.append([
            item["research_area"],
            item["count"],
        ])

    # --------------------------------------------------
    # Journals
    # --------------------------------------------------

    sheet = workbook.create_sheet("Journals")

    sheet.append([
        "Journal",
        "Publication Count",
    ])

    for item in journal_trend:
        sheet.append([
            item["journal"],
            item["count"],
        ])

    # --------------------------------------------------
    # Publications
    # --------------------------------------------------

    sheet = workbook.create_sheet("Publications")

    sheet.append([
        "Title",
        "Journal",
        "Publication Date",
        "Research Area",
        "DOI",
    ])

    for publication in publications:
        sheet.append([
            publication.title,
            publication.journal,
            publication.publication_date,
            publication.research_area,
            publication.doi,
        ])

    # --------------------------------------------------
    # Formatting
    # --------------------------------------------------

    for sheet in workbook.worksheets:

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for column in sheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 60)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    # --------------------------------------------------
    # Save
    # --------------------------------------------------

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer

def generate_funding_pdf(
    db: Session,
    user_name: str | None = None,
) -> BytesIO:
    """
    Generate a PDF report containing funding opportunities
    and funding analytics.
    """

    # --------------------------------------------------------------
    # Funding Statistics
    # --------------------------------------------------------------

    funding_statistics = (
        funding_opportunity_service.get_funding_statistics(db)
    )

    # --------------------------------------------------------------
    # Funding Analytics
    # --------------------------------------------------------------

    funding_by_agency = (
        funding_opportunity_service.get_funding_by_agency(db)
    )

    funding_by_research_area = (
        funding_opportunity_service.get_funding_by_research_area(db)
    )

    funding_by_status = (
        funding_opportunity_service.get_funding_by_status(db)
    )

    # --------------------------------------------------------------
    # Upcoming Deadlines
    # --------------------------------------------------------------

    upcoming_deadlines = (
        funding_opportunity_service.get_upcoming_deadlines(
            db=db,
            days=30,
        )
    )

    # --------------------------------------------------------------
    # Funding Opportunities
    # --------------------------------------------------------------

    funding_result = (
        funding_opportunity_service.get_funding_opportunities(
            db=db,
            page=1,
            page_size=100,
        )
    )

    funding_opportunities = funding_result["items"]

    # --------------------------------------------------------------
    # Prepare Report Data
    # --------------------------------------------------------------

    report_data = {
        "funding_statistics": funding_statistics,
        "funding_by_agency": funding_by_agency,
        "funding_by_research_area": funding_by_research_area,
        "funding_by_status": funding_by_status,
        "upcoming_deadlines": upcoming_deadlines,
        "funding_opportunities": funding_opportunities,
    }

    # --------------------------------------------------------------
    # Generate PDF
    # --------------------------------------------------------------

    pdf_generator = PDFGenerator()

    return pdf_generator.generate_funding_report(
        funding_data=report_data,
        user_name=user_name,
    )

def generate_funding_excel(
    db: Session,
    user_name: str | None = None,
) -> BytesIO:
    """
    Generate an Excel report containing funding opportunities
    and funding analytics.
    """

    funding_statistics = (
        funding_opportunity_service.get_funding_statistics(db)
    )

    funding_by_agency = (
        funding_opportunity_service.get_funding_by_agency(db)
    )

    funding_by_research_area = (
        funding_opportunity_service.get_funding_by_research_area(db)
    )

    funding_by_status = (
        funding_opportunity_service.get_funding_by_status(db)
    )

    upcoming_deadlines = (
        funding_opportunity_service.get_upcoming_deadlines(
            db=db,
            days=30,
        )
    )

    funding_result = (
        funding_opportunity_service.get_funding_opportunities(
            db=db,
            page=1,
            page_size=1000,
        )
    )

    funding_opportunities = funding_result["items"]

    workbook = Workbook()

    # --------------------------------------------------
    # Statistics
    # --------------------------------------------------

    sheet = workbook.active
    sheet.title = "Statistics"

    sheet.append([
        "Research Funding & Innovation Intelligence Platform"
    ])

    sheet.append([
        "Funding Opportunities Report"
    ])

    sheet.append([
        "User",
        user_name or "N/A",
    ])

    sheet.append([])

    sheet.append([
        "Metric",
        "Value",
    ])

    for key, value in funding_statistics.items():
        sheet.append([
            str(key).replace("_", " ").title(),
            value,
        ])

    # --------------------------------------------------
    # Agency
    # --------------------------------------------------

    sheet = workbook.create_sheet("By Agency")

    sheet.append([
        "Agency",
        "Opportunity Count",
    ])

    for item in funding_by_agency:
        sheet.append([
            item.agency,
            item.count,
        ])

    # --------------------------------------------------
    # Research Area
    # --------------------------------------------------

    sheet = workbook.create_sheet("By Research Area")

    sheet.append([
        "Research Area",
        "Opportunity Count",
    ])

    for item in funding_by_research_area:
        sheet.append([
            item.research_area,
            item.count,
        ])

    # --------------------------------------------------
    # Status
    # --------------------------------------------------

    sheet = workbook.create_sheet("By Status")

    sheet.append([
        "Status",
        "Opportunity Count",
    ])

    for item in funding_by_status:
        sheet.append([
            item.status,
            item.count,
        ])

    # --------------------------------------------------
    # Upcoming Deadlines
    # --------------------------------------------------

    sheet = workbook.create_sheet("Upcoming Deadlines")

    sheet.append([
        "Title",
        "Agency",
        "Deadline",
        "Days Remaining",
    ])

    for item in upcoming_deadlines:
        sheet.append([
            item["title"],
            item["agency"],
            item["deadline"],
            item["days_remaining"],
        ])

    # --------------------------------------------------
    # Funding Opportunities
    # --------------------------------------------------

    sheet = workbook.create_sheet("Opportunities")

    sheet.append([
        "Title",
        "Agency",
        "Research Area",
        "Funding Amount",
        "Deadline",
        "Status",
    ])

    for opportunity in funding_opportunities:
        sheet.append([
            opportunity.title,
            opportunity.agency,
            opportunity.research_area,
            opportunity.funding_amount,
            opportunity.deadline,
            opportunity.status,
        ])

    # --------------------------------------------------
    # Formatting
    # --------------------------------------------------

    for sheet in workbook.worksheets:

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for column in sheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 60)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer

def generate_patents_pdf(
    db: Session,
    user_name: str,
):
    """
    Generate a PDF report containing patent intelligence,
    technology intelligence, innovation scores, and
    commercialization recommendations.
    """

    patent_data = {
        "statistics": patent_service.get_patent_statistics(db),

        "technology_analytics": (
            patent_service.get_patents_by_technology(db)
        ),

        "status_analytics": (
            patent_service.get_patents_by_status(db)
        ),

        "country_analytics": (
            patent_service.get_patents_by_country(db)
        ),

        "filing_trend": (
            patent_service.get_patent_filing_trend(db)
        ),

        "top_inventors": (
            patent_service.get_top_inventors(db, limit=10)
        ),

        "top_assignees": (
            patent_service.get_top_assignees(db, limit=10)
        ),

        "recent_patents": (
            patent_service.get_recent_patents(db, limit=5)
        ),

        "emerging_technologies": (
            patent_service.get_emerging_technologies(db)
        ),

        "innovation_scores": (
            patent_service.get_all_innovation_scores(db)
        ),

        "commercialization_recommendations": (
            patent_service.get_all_commercialization_scores(db)
        ),

        "patents": (
            patent_service.get_patents(
                db=db,
                page=1,
                page_size=1000,
                sort_by="filing_date",
                order="desc",
            )["items"]
        ),
    }

    return pdf_generator.generate_patents_report(
        patent_data=patent_data,
        user_name=user_name,
    )

def generate_patents_excel(
    db: Session,
    user_name: str | None = None,
) -> BytesIO:
    """
    Generate an Excel report containing patent intelligence,
    technology analytics, innovation scores, and
    commercialization recommendations.
    """

    patent_statistics = patent_service.get_patent_statistics(db)

    patent_technology = (
        patent_service.get_patents_by_technology(db)
    )

    patent_status = (
        patent_service.get_patents_by_status(db)
    )

    patent_country = (
        patent_service.get_patents_by_country(db)
    )

    patent_filing_trend = (
        patent_service.get_patent_filing_trend(db)
    )

    top_inventors = (
        patent_service.get_top_inventors(db)
    )

    top_assignees = (
        patent_service.get_top_assignees(db)
    )

    recent_patents = (
        patent_service.get_recent_patents(db)
    )

    emerging_technologies = (
        patent_service.get_emerging_technologies(db)
    )

    innovation_scores = (
        patent_service.get_all_innovation_scores(db)
    )

    commercialization_recommendations = (
        patent_service.get_all_commercialization_scores(db)
    )

    workbook = Workbook()

    # --------------------------------------------------
    # Statistics
    # --------------------------------------------------

    sheet = workbook.active
    sheet.title = "Statistics"

    sheet.append([
        "Research Funding & Innovation Intelligence Platform"
    ])

    sheet.append([
        "Patent Intelligence Report"
    ])

    sheet.append([
        "User",
        user_name or "N/A",
    ])

    sheet.append([])

    sheet.append([
        "Metric",
        "Value",
    ])

    for key, value in patent_statistics.items():
        sheet.append([
            str(key).replace("_", " ").title(),
            value,
        ])

    # --------------------------------------------------
    # Technology
    # --------------------------------------------------

    sheet = workbook.create_sheet("Technology")

    sheet.append([
        "Technology Area",
        "Patent Count",
    ])

    for item in patent_technology:
        sheet.append([
            item["technology_area"],
            item["count"],
        ])

    # --------------------------------------------------
    # Status
    # --------------------------------------------------

    sheet = workbook.create_sheet("Status")

    sheet.append([
        "Status",
        "Patent Count",
    ])

    for item in patent_status:
        sheet.append([
            item["status"],
            item["count"],
        ])

    # --------------------------------------------------
    # Country
    # --------------------------------------------------

    sheet = workbook.create_sheet("Country")

    sheet.append([
        "Country",
        "Patent Count",
    ])

    for item in patent_country:
        sheet.append([
            item["country"],
            item["count"],
        ])

    # --------------------------------------------------
    # Filing Trend
    # --------------------------------------------------

    sheet = workbook.create_sheet("Filing Trend")

    sheet.append([
        "Year",
        "Patent Count",
    ])

    for item in patent_filing_trend:
        sheet.append([
            item["year"],
            item["count"],
        ])

    # --------------------------------------------------
    # Top Inventors
    # --------------------------------------------------

    sheet = workbook.create_sheet("Top Inventors")

    sheet.append([
        "Inventor",
        "Patent Count",
    ])

    for item in top_inventors:
        sheet.append([
            item["inventor"],
            item["count"],
        ])

    # --------------------------------------------------
    # Top Assignees
    # --------------------------------------------------

    sheet = workbook.create_sheet("Top Assignees")

    sheet.append([
        "Assignee",
        "Patent Count",
    ])

    for item in top_assignees:
        sheet.append([
            item["assignee"],
            item["count"],
        ])

    # --------------------------------------------------
    # Emerging Technologies
    # --------------------------------------------------

    sheet = workbook.create_sheet(
        "Emerging Technologies"
    )

    sheet.append([
        "Technology Area",
        "Patent Count",
        "Growth Score",
        "Trend",
        "Recommendation",
    ])

    for item in emerging_technologies:
        sheet.append([
            item["technology_area"],
            item["patent_count"],
            item["growth_score"],
            item["trend"],
            item["recommendation"],
        ])

    # --------------------------------------------------
    # Innovation Scores
    # --------------------------------------------------

    sheet = workbook.create_sheet(
        "Innovation Scores"
    )

    sheet.append([
        "Patent ID",
        "Title",
        "Innovation Score",
        "Innovation Level",
        "Reasons",
    ])

    for item in innovation_scores:
        sheet.append([
            item["patent_id"],
            item["title"],
            item["innovation_score"],
            item["innovation_level"],
            ", ".join(item["reasons"]),
        ])

    # --------------------------------------------------
    # Commercialization
    # --------------------------------------------------

    sheet = workbook.create_sheet(
        "Commercialization"
    )

    sheet.append([
        "Patent ID",
        "Title",
        "Commercialization Score",
        "Commercialization Level",
        "Recommended Action",
        "Reasons",
    ])

    for item in commercialization_recommendations:
        sheet.append([
            item["patent_id"],
            item["title"],
            item["commercialization_score"],
            item["commercialization_level"],
            item["recommended_action"],
            ", ".join(item["reasons"]),
        ])

    # --------------------------------------------------
    # Recent Patents
    # --------------------------------------------------

    sheet = workbook.create_sheet(
        "Recent Patents"
    )

    sheet.append([
        "Patent Number",
        "Title",
        "Technology Area",
        "Status",
        "Country",
        "Filing Date",
    ])

    for patent in recent_patents:
        sheet.append([
            patent.patent_number,
            patent.title,
            patent.technology_area,
            patent.status,
            patent.country,
            patent.filing_date,
        ])

    # --------------------------------------------------
    # Formatting
    # --------------------------------------------------

    for sheet in workbook.worksheets:

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for column in sheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 60)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer