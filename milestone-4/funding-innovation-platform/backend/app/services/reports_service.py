"""
Business logic for the Reports & Export System (Milestone 4, spec section
11: funding / patent / research trend / innovation intelligence /
commercialization reports, each exportable as PDF or Excel).

This service composes data by calling the *existing* read-only services for
each module (AnalyticsService, PatentAnalysisService, ResearchTrendService,
TechnologyIntelligenceService, InnovationScoringService,
CommercializationService) rather than re-querying the database directly —
each module stays the single source of truth for its own aggregation logic.
Report rendering (PDF via reportlab, Excel via openpyxl) is generic: every
report type is expressed as a `ReportPayload` of `ReportSection`s, so adding
a sixth report type later means writing one `_xxx_payload()` method, not a
new renderer.
"""
import io
import logging
from datetime import datetime, timezone

from sqlalchemy.orm import Session

from app.repositories.reports_repository import ReportsRepository
from app.schemas.reports import AvailableReport, ReportFormat, ReportPayload, ReportSection, ReportType
from app.services.analytics_service import AnalyticsService
from app.services.commercialization_service import CommercializationService
from app.services.innovation_scoring_service import InnovationScoringService
from app.services.patent_analysis_service import PatentAnalysisService
from app.services.research_trend_service import ResearchTrendService
from app.services.technology_intelligence_service import TechnologyIntelligenceService

logger = logging.getLogger("app.services.reports")

AVAILABLE_REPORTS = [
    AvailableReport(
        report_type=ReportType.FUNDING,
        title="Funding Report",
        description="Platform-wide funding opportunity, application, and bookmark statistics.",
    ),
    AvailableReport(
        report_type=ReportType.PATENT,
        title="Patent Report",
        description="Patent filing trend, competitor landscape, and citation analytics.",
    ),
    AvailableReport(
        report_type=ReportType.RESEARCH_TREND,
        title="Research Trend Report",
        description="Publication trend, emerging topics, research hotspots, and citation analytics.",
    ),
    AvailableReport(
        report_type=ReportType.INNOVATION_INTELLIGENCE,
        title="Innovation Intelligence Report",
        description="Innovation score leaderboard, technology maturity, and emerging technologies.",
    ),
    AvailableReport(
        report_type=ReportType.COMMERCIALIZATION,
        title="Commercialization Report",
        description="Active commercialization recommendations by type, platform-wide.",
    ),
]


class ReportsService:
    def __init__(self, db: Session):
        self.db = db
        self.reports_repo = ReportsRepository(db)

    def list_available(self) -> list[AvailableReport]:
        return AVAILABLE_REPORTS

    # ---- Payload builders (one per report type) ----

    def _funding_payload(self) -> ReportPayload:
        overview = AnalyticsService(self.db).overview()
        top_opportunities = self.reports_repo.top_funding_opportunities()
        sections = [
            ReportSection(
                heading="Summary",
                kind="kv",
                rows=[
                    ["Total Opportunities", str(overview["total_opportunities"])],
                    ["Total Applications", str(overview["total_applications"])],
                    ["Total Bookmarks", str(overview["total_bookmarks"])],
                ],
            ),
            ReportSection(
                heading="Opportunities by Status",
                kind="kv",
                rows=[[k, str(v)] for k, v in overview["opportunities_by_status"].items()],
            ),
            ReportSection(
                heading="Applications by Status",
                kind="kv",
                rows=[[k, str(v)] for k, v in overview["applications_by_status"].items()],
            ),
            ReportSection(
                heading="Top Opportunities by Views",
                kind="table",
                columns=["Title", "Status", "Views", "Deadline"],
                rows=[
                    [
                        o.title,
                        o.status.value if hasattr(o.status, "value") else str(o.status),
                        str(o.view_count),
                        o.application_deadline.isoformat() if o.application_deadline else "-",
                    ]
                    for o in top_opportunities
                ],
            ),
        ]
        return ReportPayload(
            report_type=ReportType.FUNDING,
            title="Funding Report",
            generated_at=datetime.now(timezone.utc),
            sections=sections,
        )

    def _patent_payload(self) -> ReportPayload:
        service = PatentAnalysisService(self.db)
        trend = service.trend()
        competitors = service.competitors(limit=15)
        sections = [
            ReportSection(
                heading="Patent Filing Trend",
                kind="table",
                columns=["Year", "Patent Count", "Total Citations"],
                rows=[[str(p.year), str(p.patent_count), str(p.total_citations)] for p in trend],
            ),
            ReportSection(
                heading="Top Assignees (Competitor Analysis)",
                kind="table",
                columns=["Assignee", "Patent Count", "Total Citations", "Technology Domains"],
                rows=[
                    [c.assignee, str(c.patent_count), str(c.total_citations), ", ".join(c.technology_domains)]
                    for c in competitors
                ],
            ),
        ]
        return ReportPayload(
            report_type=ReportType.PATENT,
            title="Patent Report",
            generated_at=datetime.now(timezone.utc),
            sections=sections,
        )

    def _research_trend_payload(self) -> ReportPayload:
        overview = ResearchTrendService(self.db).overview()
        sections = [
            ReportSection(
                heading="Citation Analytics",
                kind="kv",
                rows=[
                    ["Total Publications", str(overview.citation_analytics.total_publications)],
                    ["Total Citations", str(overview.citation_analytics.total_citations)],
                    ["Average Citations", str(overview.citation_analytics.average_citations)],
                    ["Max Citations", str(overview.citation_analytics.max_citations)],
                ],
            ),
            ReportSection(
                heading="Publication Trend",
                kind="table",
                columns=["Year", "Publications", "Total Citations"],
                rows=[[str(p.year), str(p.publication_count), str(p.total_citations)] for p in overview.publication_trend],
            ),
            ReportSection(
                heading="Emerging Topics",
                kind="table",
                columns=["Topic", "Recent Count", "Prior Count", "Growth Rate"],
                rows=[
                    [t.topic, str(t.recent_count), str(t.prior_count), f"{t.growth_rate:.2f}"]
                    for t in overview.emerging_topics
                ],
            ),
            ReportSection(
                heading="Research Hotspots",
                kind="table",
                columns=["Domain", "Recent Publications"],
                rows=[[h.domain, str(h.recent_publication_count)] for h in overview.research_hotspots],
            ),
        ]
        return ReportPayload(
            report_type=ReportType.RESEARCH_TREND,
            title="Research Trend Report",
            generated_at=datetime.now(timezone.utc),
            sections=sections,
        )

    def _innovation_intelligence_payload(self) -> ReportPayload:
        leaderboard = InnovationScoringService(self.db).leaderboard(limit=20)
        tech_service = TechnologyIntelligenceService(self.db)
        maturity = tech_service.maturity_breakdown()
        emerging = tech_service.emerging_technologies(limit=15)
        sections = [
            ReportSection(
                heading="Innovation Score Leaderboard",
                kind="table",
                columns=["Researcher", "Organization", "Overall Score", "Computed At"],
                rows=[
                    [e.researcher_full_name, e.organization or "-", str(e.overall_score), e.computed_at.isoformat()]
                    for e in leaderboard
                ],
            ),
            ReportSection(
                heading="Technology Maturity Breakdown",
                kind="table",
                columns=["Maturity Level", "Technology Count"],
                rows=[
                    [m.maturity_level.value if hasattr(m.maturity_level, "value") else str(m.maturity_level), str(m.technology_count)]
                    for m in maturity
                ],
            ),
            ReportSection(
                heading="Emerging Technologies",
                kind="table",
                columns=["Technology", "Tracked", "Maturity Level"],
                rows=[
                    [e.technology_name, "Yes" if e.is_tracked else "No", (e.maturity_level.value if e.maturity_level else "-")]
                    for e in emerging
                ],
            ),
        ]
        return ReportPayload(
            report_type=ReportType.INNOVATION_INTELLIGENCE,
            title="Innovation Intelligence Report",
            generated_at=datetime.now(timezone.utc),
            sections=sections,
        )

    def _commercialization_payload(self) -> ReportPayload:
        counts_by_type = self.reports_repo.commercialization_counts_by_type()
        recent = self.reports_repo.recent_commercialization_recommendations()
        sections = [
            ReportSection(
                heading="Active Recommendations by Type",
                kind="kv",
                rows=[[k, str(v)] for k, v in counts_by_type.items()],
            ),
            ReportSection(
                heading="Recent Recommendations",
                kind="table",
                columns=["Title", "Type", "Confidence", "Created"],
                rows=[
                    [
                        r.title,
                        r.recommendation_type.value if hasattr(r.recommendation_type, "value") else str(r.recommendation_type),
                        str(r.confidence_score),
                        r.created_at.isoformat(),
                    ]
                    for r in recent
                ],
            ),
        ]
        return ReportPayload(
            report_type=ReportType.COMMERCIALIZATION,
            title="Commercialization Report",
            generated_at=datetime.now(timezone.utc),
            sections=sections,
        )

    def build_payload(self, report_type: ReportType) -> ReportPayload:
        builders = {
            ReportType.FUNDING: self._funding_payload,
            ReportType.PATENT: self._patent_payload,
            ReportType.RESEARCH_TREND: self._research_trend_payload,
            ReportType.INNOVATION_INTELLIGENCE: self._innovation_intelligence_payload,
            ReportType.COMMERCIALIZATION: self._commercialization_payload,
        }
        payload = builders[report_type]()
        logger.info("Generated %s report payload with %d sections", report_type.value, len(payload.sections))
        return payload

    # ---- Renderers ----

    def render(self, report_type: ReportType, fmt: ReportFormat) -> tuple[bytes, str, str]:
        """Returns (file_bytes, filename, media_type)."""
        payload = self.build_payload(report_type)
        stamp = payload.generated_at.strftime("%Y%m%d-%H%M%S")
        if fmt == ReportFormat.PDF:
            content = self._render_pdf(payload)
            return content, f"{report_type.value}-report-{stamp}.pdf", "application/pdf"
        content = self._render_excel(payload)
        return (
            content,
            f"{report_type.value}-report-{stamp}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @staticmethod
    def _render_pdf(payload: ReportPayload) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph(payload.title, styles["Title"]),
            Paragraph(f"Generated: {payload.generated_at.strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
            Spacer(1, 0.5 * cm),
        ]

        for section in payload.sections:
            story.append(Paragraph(section.heading, styles["Heading2"]))
            rows = section.rows
            if not rows:
                story.append(Paragraph("No data available.", styles["Normal"]))
            else:
                table_data = [section.columns] + rows if section.kind == "table" and section.columns else rows
                table = Table(table_data, repeatRows=1 if (section.kind == "table" and section.columns) else 0)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                        if section.kind == "table" and section.columns
                        else [
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(table)
            story.append(Spacer(1, 0.4 * cm))

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def _render_excel(payload: ReportPayload) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for section in payload.sections:
            sheet_name = section.heading[:31] or "Sheet"
            ws = wb.create_sheet(title=sheet_name)
            ws.append([f"{payload.title} — generated {payload.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"])
            ws.append([])

            if section.kind == "table" and section.columns:
                ws.append(section.columns)
                header_row = ws.max_row
                for col_idx in range(1, len(section.columns) + 1):
                    cell = ws.cell(row=header_row, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
            for row in section.rows:
                ws.append(row)

            for col_idx, _ in enumerate(section.columns or (section.rows[0] if section.rows else []), start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22

        if not wb.sheetnames:
            wb.create_sheet(title="Report")

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
